import openpyxl

def getRowCount(file, sheetName):
 workbook = openpyxl.load_workbook(file)
 sheet = workbook.get_sheet_by_name(sheetName)
 return(sheet.max_row)

def getColumnCount(file, sheetName):
 workbook = openpyxl.load_workbook(file)
 sheet = workbook.get_sheet_by_name(sheetName)
 return(sheet.max_column)

def readData(file,sheetName, rownum, columnno):
 workbook = openpyxl.load_workbook(file)
 sheet = workbook.get_sheet_by_name(sheetName)
 return sheet.cell(row=rownum, column=columnno).value

def writeData(file,sheetName, rownum, columnno, data):
 workbook = openpyxl.load_workbook(file)
 sheet = workbook.get_sheet_by_name(sheetName)
 sheet.cell(row=rownum, column=columnno).value = data
 workbook.save(file)