'''
Data driven testing can be done on data from files like excel, or from databases, etc.
OpenPyxl Module will help us get methods to read and write data in excel.
'''
import openpyxl

path = r"C:\Users\ADMIN\Desktop\program\visual studio code\Selpy\SampleData.xlsx"

workbook = openpyxl.load_workbook(path)

# since there is only one sheet in our excel file we have to use: sheet = workbook.active
# if there were more than 1 sheet we have to get it by using: 
sheet = workbook.get_sheet_by_name("SalesOrders")

# to perform the for loop we need to know the number of rows in columns in the excel sheet
rows = sheet.max_row
columns = sheet.max_column

print('number of rows: ', rows)
print('number of columns: ', columns)

for r in range(1, rows+1):
 for c in range(1, columns+1):
  print(sheet.cell(row=r, column=c).value, end="    ")  # to capture the value from excel
 print()